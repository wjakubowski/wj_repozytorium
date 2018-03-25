#!/usr/bin/env python3
import cv2, os, re, numpy
from matplotlib import pyplot as plt
from matplotlib.image  import imread
from glob import glob
from openpyxl import Workbook
from openpyxl.compat import range
	
def clean_file(file_name):
	with open(file_name,'w') as pl:
		pass

"""def write_line(file_name,list_of_elements):
	with open(file_name,'a') as pl:
		for s in list_of_elements:
			pl.write(str(s)+'\t')
		pl.write('\n')"""
		
def txt_write(file_name,dane):
	with open(file_name,'a') as pl:
		for row in range(len(dane)):
			for col in range(len(dane[row])):
				pl.write(str(dane[row][col])+'\t')
			pl.write('\n')
		
def excel_clean_file(file_name):
	wb = Workbook()
	ws = wb.get_sheet_by_name('Sheet')
	for row in ws['A1:G1000']:
		for cell in row:
			cell.value = None
	wb.save(file_name) 
	
def excel_write(file_name,dane):
	wb = Workbook()
	ws = wb.get_sheet_by_name('Sheet')
	for row in range(len(dane)):
		for col in range(len(dane[row])):
			ws.cell(column=col+1, row=row+1, value=dane[row][col])
	wb.save(file_name) 

def procesImage(imageFile, border_gray_level):
	mainImage=imread(imageFile)
	mainGrayImage = cv2.cvtColor(mainImage, cv2.COLOR_BGR2GRAY)
	ret, mainBinarizedImage = cv2.threshold(mainGrayImage, border_gray_level ,255 ,cv2.THRESH_BINARY)

	#find excluded region file if exist
	head, tail = os.path.split(imageFile)
	excludedRegionImagePath = "{}/ExcludedRegion_{}".format(head,tail)
	excludedPixels = 0
	if os.path.isfile(excludedRegionImagePath):
		excludedRegionImage = imread(excludedRegionImagePath)
		excludedRegionGrayImage = cv2.cvtColor(excludedRegionImage, cv2.COLOR_BGR2GRAY)
		ret, excludedRegionBinarizedImage = cv2.threshold(excludedRegionGrayImage, 20 ,255 ,cv2.THRESH_BINARY)
		excludedPixels = excludedRegionBinarizedImage.sum() / 255
		excludedRegionInvertedBinarizedImage = cv2.bitwise_not(excludedRegionBinarizedImage)
		mainBinarizedImage = cv2.bitwise_and(mainBinarizedImage, excludedRegionInvertedBinarizedImage)
	
	pixels = mainBinarizedImage.shape[0] * mainBinarizedImage.shape[1] - excludedPixels
	marked_pixels = mainBinarizedImage.sum() / 255
	return pixels, marked_pixels, mainBinarizedImage, mainGrayImage
	
def procesImages(imageFiles, border_gray_level, outputName = "results"):
	try:
		os.mkdir("image_comparison")
	except (OSError):
		pass
	excel_clean_file(outputName + ".xlsx")
	clean_file(outputName + ".txt")
	str_record=[["file name", "pixels", "marked piels", "border gray level"]]
	#write_line(outputName + ".txt",["file name", "pixels", "marked piels", "border gray level"])
	
	for imageFile in imageFiles:
		pixels, marked_pixels, mainBinarizedImage, mainGrayImage = procesImage(imageFile, border_gray_level)
		
		#write_line(outputName + '.txt',[imageFile, pixels, marked_pixels, border_gray_level])
		str_record.append([imageFile, pixels, marked_pixels, border_gray_level])
		print(imageFile, pixels, marked_pixels, border_gray_level)
		
		image_comparison = numpy.concatenate((mainBinarizedImage,mainGrayImage),axis=1)
		plt.axis('off')
		file_name=os.path.split(imageFile)[1]
		plt.imsave(".\\image_comparison/"+file_name, image_comparison, cmap="gray")
	txt_write(outputName + ".txt",str_record)
	excel_write(outputName + ".xlsx",str_record)
		
if __name__ == '__main__':	
	dirpath = "."
	pattern = os.path.join(dirpath, '*.%s' % "jpg")
	imageFiles = glob(pattern)
	#procesImages(imageFiles)
	procesImages(['C:\\Users\\Wojciech\\Dropbox\\analiza_obrazu\\GUI\\pare_zdjec\\NaT_1.lif_Series036Snapshot1_ch00.jpg'])
